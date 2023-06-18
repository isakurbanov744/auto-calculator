import pandas as pd
import openpyxl
import datetime
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox


def calculate(wb, file_name, column_letter, column_num):
    # Read the first file from the "Main" sheet

    df1 = pd.read_excel(
        f'sales/{file_name}_category_sales.xlsx', sheet_name='Main')

    # Read the second file using openpyxl

    sheet = wb['MARNEULI']

    # Get the values in column C of the second file (starting from cell C5)
    names_column = [cell.value for cell in sheet[column_letter][2:]]

    # Compare names and update column F in the second file
    second_column_values = df1.values.tolist()

    for index, vals in enumerate(names_column, start=3):
        for name in second_column_values:
            if vals != None and vals in name[0]:
                sheet.cell(row=index, column=column_num).value = name[1]
                # print(index, name[0])

    # Get today's date
    today = datetime.date.today()

    # Format the date as day.month.year
    formatted_date = today.strftime("%d.%m.%Y")

    current_month = datetime.date.today().month

    today = str(formatted_date).zfill(2)

    sheet.cell(row=1, column=3).value = f'01.{current_month}.23 - {str(today)}'


def file_generator():
    col_num = [22, 6, 14]
    file_name = ['001', '002', '004']
    column_letter = 'C'
    wb = openpyxl.load_workbook('template/template.xlsx')

    for i in range(3):
        try:
            calculate(wb, file_name[i], column_letter, col_num[i])
        except:
            messagebox.showerror(
                "Error", f"Sales data couldn't be found")
            return

    wb.save('DAPNA - IYUN HEDEF 2023.xlsx')

    messagebox.showinfo(
        "Process Complete", "file generated successfully.")
